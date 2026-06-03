import os
import hashlib
import io
from datetime import date
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from dotenv import load_dotenv
from database import init_db, get_db, fetchone, fetchall, execute
import fund_logic
import openpyxl
from openpyxl.styles import Font, PatternFill

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def seed_admin():
    conn = get_db()
    row = fetchone(conn, "SELECT COUNT(*) as c FROM investisseurs")
    conn.close()
    if row and row["c"] == 0:
        conn2 = get_db()
        execute(conn2,
            "INSERT INTO investisseurs (nom, password_hash, depot_total, role) VALUES (?,?,?,?)",
            ("Malick", hash_password("admin123"), 76.0, "admin")
        )
        conn2.commit()
        conn2.close()
        print("Admin créé : Malick / admin123")


init_db()
seed_admin()


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    conn = get_db()
    row = fetchone(conn, "SELECT * FROM investisseurs WHERE id=?", (uid,))
    conn.close()
    return row


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user or user.get("role") != "admin":
            return redirect(url_for("dashboard"))
        return fn(*args, **kwargs)
    return wrapper


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        nom = request.form.get("nom", "").strip()
        password = request.form.get("password", "")
        conn = get_db()
        row = fetchone(conn, "SELECT * FROM investisseurs WHERE nom=?", (nom,))
        conn.close()
        if row and row["password_hash"] == hash_password(password):
            session["user_id"] = row["id"]
            return redirect(url_for("dashboard"))
        error = "Nom ou mot de passe incorrect."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    if user["role"] == "admin":
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("investor_dashboard"))


@app.route("/admin")
@login_required
@admin_required
def admin_dashboard():
    stats = fund_logic.get_admin_stats()
    today = date.today().isoformat()
    return render_template("dashboard_admin.html", stats=stats, today=today, user=current_user())


@app.route("/investor")
@login_required
def investor_dashboard():
    user = current_user()
    stats = fund_logic.get_investor_stats(user["id"])
    return render_template("dashboard_investor.html", stats=stats, user=user)


# ── Solde ─────────────────────────────────────────────────────────────────────

@app.route("/admin/solde", methods=["POST"])
@login_required
@admin_required
def save_solde():
    date_str = request.form.get("date", date.today().isoformat())
    try:
        solde = float(request.form.get("solde", ""))
    except ValueError:
        flash("Solde invalide.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    existing = fetchone(conn, "SELECT id FROM soldes WHERE date=?", (date_str,))
    conn.close()
    if existing:
        flash("Un solde existe déjà pour cette date.", "error")
        return redirect(url_for("admin_dashboard"))

    fund_logic.save_solde(date_str, solde)
    flash("Solde enregistré.", "success")
    return redirect(url_for("admin_dashboard"))


# ── Investisseurs ─────────────────────────────────────────────────────────────

@app.route("/admin/investisseur/ajouter", methods=["POST"])
@login_required
@admin_required
def add_investor():
    nom = request.form.get("nom", "").strip()
    password = request.form.get("password", "")
    try:
        depot = float(request.form.get("depot", 0) or 0)
    except ValueError:
        depot = 0.0

    if not nom or not password:
        flash("Nom et mot de passe requis.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    existing = fetchone(conn, "SELECT id FROM investisseurs WHERE nom=?", (nom,))
    conn.close()
    if existing:
        flash("Ce nom existe déjà.", "error")
        return redirect(url_for("admin_dashboard"))

    conn2 = get_db()
    execute(conn2,
        "INSERT INTO investisseurs (nom, password_hash, depot_total, role) VALUES (?,?,?,?)",
        (nom, hash_password(password), depot, "investor")
    )
    conn2.commit()
    conn2.close()
    flash(f"{nom} ajouté.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/investisseur/fonds", methods=["POST"])
@login_required
@admin_required
def add_funds():
    investor_id = request.form.get("investor_id", "")
    try:
        montant = float(request.form.get("montant", ""))
    except ValueError:
        flash("Montant invalide.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    row = fetchone(conn, "SELECT depot_total FROM investisseurs WHERE id=?", (investor_id,))
    conn.close()
    if not row:
        flash("Investisseur introuvable.", "error")
        return redirect(url_for("admin_dashboard"))

    conn2 = get_db()
    execute(conn2, "UPDATE investisseurs SET depot_total=? WHERE id=?",
            (row["depot_total"] + montant, investor_id))
    conn2.commit()
    conn2.close()
    flash("Fonds ajoutés.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/investisseur/retrait", methods=["POST"])
@login_required
@admin_required
def retrait_funds():
    investor_id = request.form.get("investor_id", "")
    try:
        montant = float(request.form.get("montant", ""))
    except ValueError:
        flash("Montant invalide.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    row = fetchone(conn, "SELECT depot_total, nom FROM investisseurs WHERE id=?", (investor_id,))
    conn.close()
    if not row:
        flash("Investisseur introuvable.", "error")
        return redirect(url_for("admin_dashboard"))

    nouveau_depot = round(row["depot_total"] - montant, 4)
    if nouveau_depot < 0:
        flash(f"Retrait impossible : solde insuffisant ({row['depot_total']} USDT disponible).", "error")
        return redirect(url_for("admin_dashboard"))

    conn2 = get_db()
    execute(conn2, "UPDATE investisseurs SET depot_total=? WHERE id=?", (nouveau_depot, investor_id))
    conn2.commit()
    conn2.close()
    flash(f"Retrait de {montant} USDT effectué pour {row['nom']}.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/investisseur/modifier", methods=["POST"])
@login_required
@admin_required
def edit_investor():
    investor_id = request.form.get("investor_id", "")
    try:
        nouveau_depot = float(request.form.get("depot_total", ""))
    except ValueError:
        flash("Montant invalide.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    execute(conn, "UPDATE investisseurs SET depot_total=? WHERE id=?", (nouveau_depot, investor_id))
    conn.commit()
    conn.close()
    flash("Dépôt mis à jour.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/investisseur/role", methods=["POST"])
@login_required
@admin_required
def change_role():
    investor_id = request.form.get("investor_id", "")
    new_role = request.form.get("role", "investor")
    if new_role not in ("admin", "investor"):
        flash("Rôle invalide.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    execute(conn, "UPDATE investisseurs SET role=? WHERE id=?", (new_role, investor_id))
    conn.commit()
    conn.close()
    flash("Rôle mis à jour.", "success")
    return redirect(url_for("admin_dashboard"))


# ── Export Excel ──────────────────────────────────────────────────────────────

def style_header(ws, headers):
    hf = PatternFill("solid", fgColor="0d1117")
    hfont = Font(bold=True, color="00e87a")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hf
        ws.column_dimensions[c.column_letter].width = max(len(h) + 6, 16)


@app.route("/admin/export")
@login_required
@admin_required
def export_admin():
    conn = get_db()
    soldes = fetchall(conn, "SELECT * FROM soldes ORDER BY date DESC")
    investors = fetchall(conn, "SELECT * FROM investisseurs")
    gains = fetchall(conn, "SELECT * FROM gains_investisseurs ORDER BY date DESC")
    conn.close()

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Soldes"
    style_header(ws1, ["Date", "Solde (USDT)", "Gain jour", "% jour", "Gain total"])
    for s in soldes:
        ws1.append([s["date"], s["solde"], s["gain_jour"], s["pct_jour"], s["gain_total"]])

    ws2 = wb.create_sheet("Investisseurs")
    style_header(ws2, ["Nom", "Dépôt (USDT)", "Rôle", "Membre depuis"])
    for inv in investors:
        ws2.append([inv["nom"], inv["depot_total"], inv["role"], str(inv.get("date_entree", ""))])

    ws3 = wb.create_sheet("Gains par investisseur")
    style_header(ws3, ["Date", "Investisseur", "Part %", "Gain (USDT)"])
    for g in gains:
        ws3.append([g["date"], g["investisseur_nom"], g["part_pct"], g["gain"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="MO_Capital_historique.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/investor/export")
@login_required
def export_investor():
    user = current_user()
    conn = get_db()
    gains = fetchall(conn,
        "SELECT * FROM gains_investisseurs WHERE investisseur_id=? ORDER BY date DESC",
        (user["id"],)
    )
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mes gains"
    style_header(ws, ["Date", "Part %", "Gain (USDT)"])
    for g in gains:
        ws.append([g["date"], g["part_pct"], g["gain"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=f"MO_Capital_{user['nom']}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
